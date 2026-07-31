"""
Google Sheet integration.

- profile tabs fed by the bots
- an archive tab (import of a spreadsheet WITH its colors)
- Outcome column = dropdown, automatic row coloring
- manual edits (Outcome, Notes) are preserved on every sync
- auto "No reply (3+ weeks)" a configurable number of days after applying
"""

from datetime import date
from pathlib import Path
import gspread
from google.oauth2.service_account import Credentials

import appconfig
import db

SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive.file"]

GSPREAD_TIMEOUT = 30  # seconds: a Google stall must not freeze the bot

HEADERS = ["Added", "Company", "Title", "Location", "Start date", "Score",
           "Link", "Source", "Status", "Applied on", "Outcome", "Notes"]
NCOLS = len(HEADERS)
COL_LINK = 6        # 0-based index of "Link"
COL_OUTCOME = 10    # 0-based index of "Outcome" (column K)

STATUS = {"kept": "To apply", "applied": "Applied"}

# Outcome dropdown options
OUTCOME_OPTIONS = ["No reply (3+ weeks)", "Rejected", "Interview", "Accepted", "In progress"]


def client(cfg):
    gs = cfg["google_sheet"]
    creds = Credentials.from_service_account_file(
        str(appconfig.ROOT / gs["credentials_file"]), scopes=SCOPES)
    gc = gspread.authorize(creds)
    gc.set_timeout(GSPREAD_TIMEOUT)  # a Google stall must not freeze the bot
    return gc.open_by_key(gs["spreadsheet_id"])


def ensure_ws(sh, title, rows=200, cols=NCOLS):
    try:
        return sh.worksheet(title)
    except gspread.WorksheetNotFound:
        return sh.add_worksheet(title=title, rows=rows, cols=cols)


def _read_manual(ws):
    """Existing rows in the tab: {link: row} + non-empty rows without a link.
    Everything is preserved, so a manual edit is never overwritten."""
    by_url, keyless = {}, []
    try:
        vals = ws.get_all_values()
    except Exception:
        return by_url, keyless
    for r in vals[1:]:
        if len(r) > COL_LINK and r[COL_LINK]:
            by_url[r[COL_LINK]] = r
        elif any((c or "").strip() for c in r):
            keyless.append(r)  # manual row without a link (e.g. spontaneous application)
    return by_url, keyless


def _outcome_notes(row):
    """Outcome + Notes of an existing row (list of cells)."""
    res = row[COL_OUTCOME] if len(row) > COL_OUTCOME else ""
    notes = row[COL_OUTCOME + 1] if len(row) > COL_OUTCOME + 1 else ""
    return res, notes


def _rows_from_db(conn, profil, manual, keyless, no_reply_days):
    today = date.today()
    out = []
    db_urls = set()
    for r in db.list_for_sheet(conn, profil):
        url = r["url"]
        db_urls.add(url)
        res, notes = _outcome_notes(manual.get(url, []))
        # auto "No reply (3+ weeks)" if applied more than N days ago and outcome empty
        if not res and r["queue_status"] == "applied" and r["applied_at"]:
            try:
                days = (today - date.fromisoformat(r["applied_at"][:10])).days
                if days >= no_reply_days:
                    res = "No reply (3+ weeks)"
            except ValueError:
                pass
        out.append([
            (r["found_at"] or "")[:10],
            r["entreprise"], r["titre"], r["lieu"],
            (r["date_debut"] or ""), r["score"], url, r["source"],
            STATUS.get(r["queue_status"], r["queue_status"]),
            (r["applied_at"] or ""), res, notes,
        ])
    # manual rows added in the tab (URL not in the DB): preserved
    for url, row in manual.items():
        if url not in db_urls:
            out.append((list(row) + [""] * NCOLS)[:NCOLS])
    # manual rows without a link: preserved too
    for row in keyless:
        out.append((list(row) + [""] * NCOLS)[:NCOLS])
    return out


def _dropdown_request(ws):
    return {"setDataValidation": {
        "range": {"sheetId": ws.id, "startRowIndex": 1,
                  "startColumnIndex": COL_OUTCOME, "endColumnIndex": COL_OUTCOME + 1},
        "rule": {
            "condition": {"type": "ONE_OF_LIST",
                          "values": [{"userEnteredValue": v} for v in OUTCOME_OPTIONS]},
            "showCustomUi": True, "strict": False,
        }}}


def _color_rules(ws, sep=";"):
    """Whole row colored by Outcome (col K): red = dead, green = positive."""
    rng = {"sheetId": ws.id, "startRowIndex": 1, "startColumnIndex": 0,
           "endColumnIndex": NCOLS}
    def rule(regex, r, g, b):
        return {"addConditionalFormatRule": {"index": 0, "rule": {
            "ranges": [rng],
            "booleanRule": {
                "condition": {"type": "CUSTOM_FORMULA", "values": [
                    {"userEnteredValue": f'=REGEXMATCH($K2{sep}"{regex}")'}]},
                "format": {"backgroundColor": {"red": r, "green": g, "blue": b}},
            }}}}
    return [
        rule("(?i)reject|no reply", 0.96, 0.80, 0.80),      # red = dead
        rule("(?i)interview|accept", 0.85, 0.92, 0.83),     # green = positive
    ]


def sync_values(sh, conn, cfg, profil):
    """Rewrite values while preserving Outcome/Notes; set the dropdown."""
    tab = cfg["profiles"][profil]["tab"]
    no_reply_days = cfg.get("behavior", {}).get("no_reply_days", 21)
    ws = ensure_ws(sh, tab)
    manual, keyless = _read_manual(ws)
    data = [HEADERS] + _rows_from_db(conn, profil, manual, keyless, no_reply_days)
    ws.clear()
    ws.update(values=data, range_name="A1", value_input_option="RAW")
    ws.format("A1:L1", {"textFormat": {"bold": True,
              "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
              "backgroundColor": {"red": 0.12, "green": 0.16, "blue": 0.27},
              "horizontalAlignment": "CENTER"})
    ws.freeze(rows=1)
    try:
        sh.batch_update({"requests": [_dropdown_request(ws)]})
    except Exception as e:  # noqa: BLE001
        print("  (dropdown skipped:", e, ")")
    return ws


def push_profile(sh, conn, cfg, profil):
    """Full setup of a tab: values + dropdown + color rules (once)."""
    ws = sync_values(sh, conn, cfg, profil)
    locale = (sh.fetch_sheet_metadata().get("properties", {})
              .get("locale", "en_US"))
    sep = ";" if str(locale).lower().startswith("fr") else ","
    try:
        sh.batch_update({"requests": _color_rules(ws, sep)})
    except Exception as e:  # noqa: BLE001
        print("  (conditional colors skipped:", e, ")")
    return ws


def quick_push(profil):
    """For the bot: connect + rewrite the profile's tab."""
    cfg = appconfig.load_config()
    conn = db.connect()
    sh = client(cfg)
    return sync_values(sh, conn, cfg, profil)


# --------------------------------------------------------------------------
# Import of an archive spreadsheet, keeping its colors
# --------------------------------------------------------------------------
def _excel_fill(cell):
    """Background color of an Excel cell -> (r,g,b) 0..1, or None."""
    if cell.fill is None or cell.fill.patternType != "solid":
        return None
    fg = cell.fill.fgColor
    if fg is None:
        return None
    if fg.type == "rgb" and isinstance(fg.rgb, str) and fg.rgb not in ("00000000", "FFFFFFFF"):
        h = fg.rgb[-6:]
        return (int(h[0:2], 16) / 255, int(h[2:4], 16) / 255, int(h[4:6], 16) / 255)
    if fg.type == "theme":
        # approximations (default Office theme)
        return {9: (0.80, 0.90, 0.75), 6: (0.80, 0.90, 0.75),
                5: (0.99, 0.85, 0.70), 4: (0.99, 0.85, 0.70)}.get(fg.theme)
    return None


def import_archive(sh, cfg, xlsx_path: Path):
    from openpyxl import load_workbook
    tab = cfg["google_sheet"]["archive_tab"]
    wb = load_workbook(xlsx_path)
    src = wb.active

    values, colors = [], []
    for row in src.iter_rows():
        values.append([("" if c.value is None else str(c.value)) for c in row])
        colors.append(_excel_fill(row[0]))  # color = background of column A

    ncols = max((len(r) for r in values), default=10)
    ws = ensure_ws(sh, tab, rows=max(len(values) + 5, 50), cols=max(ncols, 10))
    ws.clear()
    if not values:
        return ws, 0
    ws.update(values=values, range_name="A1", value_input_option="RAW")
    ws.format(f"A1:{chr(64 + ncols)}1", {"textFormat": {"bold": True}})
    ws.freeze(rows=1)

    # row colors (batched to limit requests)
    reqs = []
    for i, col in enumerate(colors):
        if not col or i == 0:  # skip the header
            continue
        r, g, b = col
        reqs.append({"repeatCell": {
            "range": {"sheetId": ws.id, "startRowIndex": i, "endRowIndex": i + 1,
                      "startColumnIndex": 0, "endColumnIndex": ncols},
            "cell": {"userEnteredFormat": {"backgroundColor": {"red": r, "green": g, "blue": b}}},
            "fields": "userEnteredFormat.backgroundColor"}})
    for k in range(0, len(reqs), 200):
        sh.batch_update({"requests": reqs[k:k + 200]})
    return ws, len(values) - 1
