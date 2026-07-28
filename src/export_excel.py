"""
Local Excel tracker, rebuilt from the DB.

ONE living file (Suivi_<profile>.xlsx), always up to date, never 40 versions.
Row color by Outcome:
  Rejected -> red, Interview/positive -> green, No reply -> grey.
"""

import os
from pathlib import Path
import db
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent

HEADERS = ["Added", "Company", "Title", "Location", "Start date", "Score",
           "Link", "Source", "Status", "Applied on", "Outcome", "Notes"]
COL_LINK = 6
COL_OUTCOME = 10

STATUS = {"kept": "To apply", "applied": "Applied"}

FILL_RED = PatternFill("solid", fgColor="F4CCCC")
FILL_GREEN = PatternFill("solid", fgColor="D9EAD3")
FILL_HEADER = PatternFill("solid", fgColor="1F2A44")


def _read_existing(path: Path):
    """Read manual edits from an existing xlsx: {link: full_row} + rows without a
    link. So nothing is overwritten on the next export."""
    by_url, keyless = {}, []
    if not path.exists():
        return by_url, keyless
    try:
        ws = load_workbook(path).active
    except Exception:
        return by_url, keyless
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = ["" if c is None else str(c) for c in row]
        if len(cells) > COL_LINK and cells[COL_LINK]:
            by_url[cells[COL_LINK]] = cells   # full row (not just outcome/notes)
        elif any(c.strip() for c in cells):
            keyless.append(cells)
    return by_url, keyless


def _fill_for(outcome: str):
    r = (outcome or "").lower()
    if "reject" in r or "no reply" in r:      # dead
        return FILL_RED
    if "interview" in r or "accept" in r:     # positive
        return FILL_GREEN
    return None


def _append_colored(ws, row):
    ws.append(row)
    fill = _fill_for(row[COL_OUTCOME] if len(row) > COL_OUTCOME else "")
    if fill:
        for c in ws[ws.max_row]:
            c.fill = fill


def export(conn, profil: str, filename: str) -> Path:
    path = ROOT / filename
    by_url, keyless = _read_existing(path)   # preserve manual edits
    rows = db.list_for_sheet(conn, profil)

    wb = Workbook()
    ws = wb.active
    ws.title = profil.capitalize()
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center")

    def _outcome_notes(cells):
        res = cells[COL_OUTCOME] if len(cells) > COL_OUTCOME else ""
        notes = cells[COL_OUTCOME + 1] if len(cells) > COL_OUTCOME + 1 else ""
        return res, notes

    db_urls = set()
    for r in rows:
        url = r["url"]
        db_urls.add(url)
        res, notes = _outcome_notes(by_url.get(url, []))
        _append_colored(ws, [
            (r["found_at"] or "")[:10], r["entreprise"], r["titre"], r["lieu"],
            (r["date_debut"] or ""), r["score"], url, r["source"],
            STATUS.get(r["queue_status"], r["queue_status"]),
            (r["applied_at"] or ""), res, notes,
        ])
    # manual rows (link not in the DB, or no link): preserved in full
    for url, cells in by_url.items():
        if url not in db_urls:
            _append_colored(ws, (list(cells) + [""] * len(HEADERS))[:len(HEADERS)])
    for cells in keyless:
        _append_colored(ws, (list(cells) + [""] * len(HEADERS))[:len(HEADERS)])

    widths = [12, 22, 46, 24, 12, 7, 40, 16, 12, 14, 16, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    tmp = path.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    os.replace(tmp, path)   # atomic write
    return path


if __name__ == "__main__":
    c = db.connect()
    p = export(c, "alternance", "tracker_alternance.xlsx")
    print("written:", p, "|", len(db.list_for_sheet(c, "alternance")), "offers")
